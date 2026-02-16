"""
再出品シートの品番からテンプレート形式のデータを生成
- 再出品シートの「済」列が空白の品番を取得
- 画像ファイル名をテンプレートの商品画像名_1～20に設定
- CSVから商品データを取得してマージ
"""
import os
import glob
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# 設定
EXCEL_FILE = r"C:\Users\progr\Desktop\Python\mercari_dorekai\downloads\ドレ買い.xlsx"
IMAGE_DIR = r"\\LS210DNBD82\share\平良\Python\mercari_dorekai\mercari_images\re"
CSV_DIR = r"C:\Users\progr\Desktop\Python\mercari_dorekai\downloads"
CSV_PATTERN = "product_data_*.csv"

def find_latest_csv():
    """最新のproduct_data_*.csvを探す"""
    pattern = os.path.join(CSV_DIR, CSV_PATTERN)
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def get_image_files(product_number):
    """指定した品番の画像ファイル名を取得（最大20枚）"""
    if not os.path.exists(IMAGE_DIR):
        return []
    
    images = []
    for filename in os.listdir(IMAGE_DIR):
        # 品番-連番.拡張子 の形式をチェック
        if filename.startswith(f"{product_number}-"):
            parts = filename.rsplit('.', 1)
            if len(parts) == 2:
                name_part = parts[0]
                try:
                    seq = int(name_part.split('-')[1])
                    images.append((seq, filename))
                except (ValueError, IndexError):
                    continue
    
    # 連番順にソート
    images.sort(key=lambda x: x[0])
    return [filename for _, filename in images[:20]]  # 最大20枚

def read_unpublished_products(workbook):
    """再出品シートから未出品の品番と必要なデータを取得"""
    if "再出品" not in workbook.sheetnames:
        print("❌ 「再出品」シートが見つかりません")
        return {}
    
    sheet = workbook["再出品"]
    
    # シート1から販売単価データを読み込む
    price_dict = {}
    if "シート1" in workbook.sheetnames:
        sheet1 = workbook["シート1"]
        # シート1のヘッダーを確認（3行目がヘッダー）
        sheet1_headers = {}
        for col_idx, cell in enumerate(sheet1[3], start=1):
            if cell.value:
                sheet1_headers[str(cell.value).strip()] = col_idx
        
        # 品番と販売単価の列を探す
        hinban_col = None
        price_col = None
        for key, idx in sheet1_headers.items():
            if key in ["品番", "商品番号", "ヒンバン"]:
                hinban_col = idx
            if key in ["販売価格", "販売単価", "単価", "価格"]:
                price_col = idx
        
        if hinban_col and price_col:
            # 品番と価格のマッピングを作成（4行目からデータ開始）
            for row in sheet1.iter_rows(min_row=4, values_only=True):
                if len(row) >= max(hinban_col, price_col):
                    hinban = row[hinban_col - 1]
                    price = row[price_col - 1]
                    if hinban and price is not None:
                        hinban_str = str(hinban).strip().lstrip('0')
                        if hinban_str:
                            try:
                                price_dict[hinban_str] = int(float(price))
                            except (ValueError, TypeError):
                                pass
            print(f"✓ シート1から{len(price_dict)}件の価格データを読み込みました")
        else:
            print(f"⚠️ シート1に品番列または販売単価列が見つかりません")
            print(f"  利用可能な列: {list(sheet1_headers.keys())}")
    else:
        print(f"⚠️ シート1が見つかりません")
    
    # ヘッダー行を探す - 必要な列をすべて検索
    columns = {}
    col_map = {
        "品番": ["品番", "商品番号", "ヒンバン"],
        "済": ["済", "済み", "完了"],
        "ブランド名": ["ブランド名", "ブランド"],
        "丈": ["丈"],
        "商品名1": ["商品名1"],
        "商品名2": ["商品名2"],
        "サイズ": ["＊サイズ", "サイズ"]
    }
    
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=False):
        for col_idx, cell in enumerate(row, start=1):
            value = str(cell.value or "").strip()
            for key, patterns in col_map.items():
                if key not in columns and value in patterns:
                    columns[key] = col_idx
        if all(k in columns for k in ["品番", "済"]):
            break
    
    if "品番" not in columns:
        print("❌ 品番列が見つかりません")
        return {}
    
    if "済" not in columns:
        print("⚠️ 済列が見つかりません")
        return {}
    
    # 今日の日付（YYMMDD形式）
    today = datetime.now().strftime("%y%m%d")
    print(f"✓ 今日の日付: {today}")
    
    # 未出品の品番とデータを収集（済列が今日の日付の場合）
    unpublished = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        product_value = row[columns["品番"] - 1].value if len(row) >= columns["品番"] else None
        done_value = row[columns["済"] - 1].value if len(row) >= columns["済"] else None
        
        # 済列が今日の日付の場合のみ処理
        if product_value and done_value is not None:
            done_str = str(done_value).strip()
            if done_str == today:
                product_number = str(product_value).strip().lstrip('0')
                original_product_number = str(product_value).strip()
                if product_number:
                    data = {"販売単価": None}
                    
                    # シート1から価格を取得
                    if product_number in price_dict:
                        data["販売単価"] = price_dict[product_number]
                    
                    # その他の列を取得
                    data["品番_元"] = original_product_number
                    for key in ["ブランド名", "丈", "商品名1", "商品名2", "サイズ"]:
                        if key in columns and len(row) >= columns[key]:
                            val = row[columns[key] - 1].value
                            data[key] = str(val).strip() if val else ""
                        else:
                            data[key] = ""
                    
                    unpublished[product_number] = data
    
    return unpublished

def read_csv_data(csv_path):
    """
    CSVから商品データを読み込み
    条件：商品名と商品説明の最初の数字が一致する商品のみ
    戻り値：{品番: CSVデータ} の辞書
    """
    products = {}
    
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    
    for enc in encodings:
        try:
            with open(csv_path, 'r', encoding=enc) as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    # 商品名と商品説明を取得
                    product_name = (row.get('商品名', '') or '').strip()
                    product_desc = (row.get('商品説明', '') or '').strip()
                    
                    if not product_name or not product_desc:
                        continue
                    
                    # 商品名の最初の数字を抽出
                    import re
                    name_match = re.search(r'^(\d+)', product_name)
                    # 商品説明の最初の数字を抽出
                    desc_match = re.search(r'^(\d+)', product_desc)
                    
                    if name_match and desc_match:
                        name_number = name_match.group(1)
                        desc_number = desc_match.group(1)
                        
                        # 商品名と商品説明の最初の数字が一致する場合のみ
                        if name_number == desc_number:
                            # 先頭の0を除去した品番をキーにする
                            product_number = name_number.lstrip('0')
                            if product_number:
                                products[product_number] = row
            
            print(f"✓ CSV読み込み成功: {os.path.basename(csv_path)} ({enc})")
            print(f"  品番一致商品: {len(products)}件")
            return products
        except Exception as e:
            continue
    
    print(f"⚠️ CSVの読み込みに失敗しました")
    return {}

def create_template_rows(unpublished_products, csv_data, template_headers):
    """
    テンプレート形式の行データを作成
    unpublished_products: {品番: {販売単価, ブランド名, 丈, ...}} の辞書
    """
    rows = []
    
    for product_number, product_data in unpublished_products.items():
        # 画像ファイルを取得
        image_files = get_image_files(product_number)
        
        if not image_files:
            print(f"⚠️ 品番 {product_number}: 画像が見つかりません")
            continue
        
        # CSVデータを取得
        csv_row = csv_data.get(product_number)
        
        if not csv_row:
            print(f"⚠️ 品番 {product_number}: CSVデータが見つかりません（商品名と説明の品番不一致）")
            continue
        
        # テンプレート形式の行を作成
        row_data = {}
        
        # 商品画像名_1～20を設定
        for i in range(1, 21):
            col_name = f"商品画像名_{i}"
            if i <= len(image_files):
                row_data[col_name] = image_files[i - 1]
            else:
                row_data[col_name] = ""
        
        # 商品名を再出品シートのデータから組み立て
        # 品番(先頭0除く) + ブランド名 + キャバクラ + ドレス + 丈 + 商品名1 + 商品名2 + キャバドレス + サイズ
        name_parts = []
        
        # 品番（先頭0除去済み）
        name_parts.append(product_number)
        
        # ブランド名（除外リスト以外）
        exclude_brands = ["SHEIN", "シーイン", "SHEIN シーイン", "ノーブランド"]
        brand_name = product_data.get("ブランド名", "").strip()
        if brand_name and brand_name not in exclude_brands:
            name_parts.append(brand_name)
        
        # キャバクラ（固定）
        name_parts.append("キャバクラ")
        
        # ドレス（固定）
        name_parts.append("ドレス")
        
        # 丈（ロングの場合は「ロングドレス」とする）
        length = product_data.get("丈", "").strip()
        if length:
            if length == "ロング":
                name_parts.append("ロングドレス")
            else:
                name_parts.append(length)
        
        # 商品名1
        if product_data.get("商品名1"):
            name_parts.append(product_data["商品名1"])
        
        # 商品名2
        if product_data.get("商品名2"):
            name_parts.append(product_data["商品名2"])
        
        # キャバドレス（固定）
        name_parts.append("キャバドレス")
        
        # サイズ（数字の場合は「数字+サイズ」）
        size = product_data.get("サイズ", "").strip()
        if size:
            if size.isdigit():
                name_parts.append(f"{size}サイズ")
            else:
                name_parts.append(size)
        
        row_data["商品名"] = " ".join(name_parts)
        row_data["商品説明"] = csv_row.get("商品説明", "")
        
        # 再出品シートの販売単価を販売価格に設定
        sales_price = product_data.get("販売単価")
        if sales_price is not None:
            row_data["販売価格"] = sales_price
        else:
            row_data["販売価格"] = ""
        
        # CSVから取得する項目
        row_data["ブランドID"] = csv_row.get("ブランドID", "")
        row_data["カテゴリID"] = csv_row.get("カテゴリID", "")
        row_data["商品の状態"] = csv_row.get("商品の状態", "")
        row_data["配送方法"] = csv_row.get("配送方法", "")
        row_data["発送元の地域"] = csv_row.get("発送元の地域", "")
        row_data["発送までの日数"] = csv_row.get("発送までの日数", "")
        row_data["商品ステータス"] = csv_row.get("商品ステータス", "")
        row_data["配送料の負担"] = csv_row.get("配送料の負担", "")
        row_data["送料ID"] = csv_row.get("送料ID", "")
        row_data["メルカリBiz配送_クール区分"] = csv_row.get("メルカリBiz配送_クール区分", "")
        
        # 在庫数を設定
        row_data["SKU1_在庫数"] = 1
        
        rows.append((product_number, row_data))
        price_display = sales_price if sales_price is not None else "未設定"
        print(f"✓ 品番 {product_number}: {len(image_files)}枚の画像, 販売価格: {price_display}")
        print(f"  商品名: {row_data['商品名']}")
    
    return rows

def write_to_excel_template(workbook, template_headers, rows):
    """
    Excelの「テンプレ」シートにデータを書き込む
    既存のデータをクリアして、2行目から新しいデータを書き込む
    """
    template_sheet = workbook["テンプレ"]
    
    # 既存データをクリア（ヘッダー行は残す）
    max_row = template_sheet.max_row
    if max_row > 1:
        template_sheet.delete_rows(2, max_row - 1)
        print(f"✓ テンプレシートの既存データをクリアしました")
    
    # データを2行目から書き込み
    for idx, (product_number, row_data) in enumerate(rows, start=2):
        for col_idx, header in enumerate(template_headers, start=1):
            value = row_data.get(header, "")
            template_sheet.cell(row=idx, column=col_idx, value=value)
    
    print(f"✓ テンプレシートに{len(rows)}件のデータを書き込みました")
    return True

def generate_template_data():
    """メイン処理"""
    print("=" * 60)
    print("テンプレートデータ生成")
    print("=" * 60)
    
    # Excelを開く
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excelファイルが見つかりません: {EXCEL_FILE}")
        return
    
    # データ読み取り用（data_only=True で数式の計算結果を取得）
    workbook_read = load_workbook(EXCEL_FILE, data_only=True)
    
    # データ書き込み用（data_only=False で数式を保持）
    workbook_write = load_workbook(EXCEL_FILE, data_only=False)
    
    # テンプレシートからヘッダーを取得
    if "テンプレ" not in workbook_write.sheetnames:
        print(f"❌ 「テンプレ」シートが見つかりません")
        workbook_read.close()
        workbook_write.close()
        return
    
    template_sheet = workbook_write["テンプレ"]
    template_headers = []
    for col in range(1, template_sheet.max_column + 1):
        header = template_sheet.cell(row=1, column=col).value
        if header:
            template_headers.append(header)
    
    print(f"✓ テンプレートヘッダー: {len(template_headers)}列")
    
    # 再出品シートから未出品の品番を取得（読み取り用ワークブック使用）
    unpublished = read_unpublished_products(workbook_read)
    if not unpublished:
        print("\n❌ 未出品の品番がありません")
        workbook_read.close()
        workbook_write.close()
        return
    
    print(f"✓ 未出品の品番: {len(unpublished)}件")
    for i, (pn, data) in enumerate(list(unpublished.items())[:10], 1):
        price_str = f" (販売単価: {data.get('販売単価')})" if data.get('販売単価') else ""
        print(f"  {i}. {pn}{price_str}")
    if len(unpublished) > 10:
        print(f"  ... 他 {len(unpublished) - 10}件")
    
    # 読み取り用ワークブックを閉じる
    workbook_read.close()
    
    # 最新のCSVを読み込み
    latest_csv = find_latest_csv()
    if not latest_csv:
        print(f"\n❌ CSVファイルが見つかりません: {CSV_PATTERN}")
        workbook_write.close()
        return
    
    print(f"\n✓ CSVファイル: {os.path.basename(latest_csv)}")
    csv_data = read_csv_data(latest_csv)
    
    # テンプレート行を生成
    print(f"\n📝 データ生成中...")
    rows = create_template_rows(unpublished, csv_data, template_headers)
    
    if not rows:
        print("\n❌ 生成できるデータがありません")
        workbook_write.close()
        return
    
    # CSVファイルとして保存
    output_file = os.path.join(CSV_DIR, f"生成_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    
    with open(output_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=template_headers)
        writer.writeheader()
        
        for product_number, row_data in rows:
            # ヘッダーに合わせてデータを整形
            csv_row = {header: row_data.get(header, "") for header in template_headers}
            writer.writerow(csv_row)
    
    print(f"\n✅ CSV保存完了: {output_file}")
    
    # Excelの「テンプレ」シートにも書き込み（書き込み用ワークブック使用）
    print(f"\n📝 Excelのテンプレシートに書き込み中...")
    if write_to_excel_template(workbook_write, template_headers, rows):
        try:
            workbook_write.save(EXCEL_FILE)
            print(f"✅ Excel保存完了: {EXCEL_FILE}")
        except Exception as e:
            print(f"❌ Excel保存エラー: {e}")
    
    workbook_write.close()
    
    print(f"\n{'='*60}")
    print(f"✅ 生成完了")
    print(f"  生成件数: {len(rows)}件")
    print(f"  CSV保存先: {output_file}")
    print(f"  Excel保存先: {EXCEL_FILE} (テンプレシート)")
    print(f"{'='*60}")

if __name__ == "__main__":
    generate_template_data()
