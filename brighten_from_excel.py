"""
Excelファイルの「再出品」シートから品番を読み取り、
画像をreフォルダに移動して1枚目を明るくするスクリプト
"""
import os
import sys
import glob
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# brighten_images.pyの関数を使用
from brighten_images import process_product_numbers

EXCEL_FILE = r"C:\Users\progr\Desktop\Python\mercari_dorekai\downloads\ドレ買い.xlsx"
SHEET_NAME = "再出品"
CSV_DIR = r"C:\Users\progr\Desktop\Python\mercari_dorekai\downloads"

def extract_product_number_from_text(text):
    """
    テキストから先頭の数字を抽出（先頭の0を除く）
    """
    if not text or not isinstance(text, str):
        return None
    
    # 先頭の数字を抽出
    match = re.match(r'^\s*(\d+)', text)
    if match:
        number = match.group(1).lstrip('0')
        return number if number else None
    return None

def extract_size_from_description(description):
    """
    商品説明からサイズを抽出
    - 数字の場合はF
    - 何も書いてない場合もF
    - 有効なサイズ値のみ抽出
    """
    if not description or not isinstance(description, str):
        return 'F'
    
    # 有効なサイズ値のリスト
    valid_sizes = [
        'XXS', 'XS', 'S', 'M', 'L', 'XL', 'XXL', '2XL', '3XL',
        'F', 'FREE', 'フリー', 'フリーサイズ',
        'S-M', 'M-L', 'L-XL',
        # EU・UK・US サイズ
        'EU32', 'EU34', 'EU36', 'EU38', 'EU40', 'EU42',
        'UK4', 'UK6', 'UK8', 'UK10', 'UK12', 'UK14',
        'US0', 'US2', 'US4', 'US6', 'US8', 'US10',
        # その他
        '0', '00', '2', '4', '6', '8', '9', '10', '11', '12', '13', '15',
        '0サイズ', '2サイズ', '4サイズ', '6サイズ', '7サイズ', '9サイズ', '11サイズ', '13サイズ', '15サイズ', '16サイズ', '36サイズ', '38サイズ', '40サイズ',
        '7号', '9号', '11号', '13号', '15号',
        # キッズサイズ
        '4y', '5y', '6y', '7y', '8y', '9y', '10y', '11y', '12y', '13y', '14y', '15y', '16y'
    ]
    
    # ＊サイズの後の行を探す
    match = re.search(r'＊サイズ\s*\n+\s*(\S+)', description)
    if not match:
        return 'F'
    
    size_value = match.group(1).strip()
    
    # 肩幅や身幅などの採寸が来た場合は空と判断
    if '幅' in size_value or 'cm' in size_value or '約' in size_value or '着丈' in size_value:
        return 'F'
    
    # 長すぎる場合はF（説明文など）
    if len(size_value) > 20:
        return 'F'
    
    # 数字のみ（サイズ表記なし）の場合
    if size_value.isdigit():
        # 1桁または2桁の数字の場合はF
        if len(size_value) <= 2:
            return 'F'
    
    # 有効なサイズリストにあればそのまま返す
    size_upper = size_value.upper()
    for valid_size in valid_sizes:
        if valid_size.upper() == size_upper:
            return size_value
    
    # 「表記」「タグ」などはF
    if '表記' in size_value or 'タグ' in size_value or '画像' in size_value:
        return 'F'
    
    # それ以外はF
    return 'F'

def clean_product_name(name):
    """
    商品名をクリーンアップ
    - スペース区切り
    - 始部分の数字を削除
    - キャバクラ、ドレス、ロングドレス、キャバドレスを削除（単語として独立している場合のみ）
    """
    if not name or not isinstance(name, str):
        return ''
    
    # 先頭の数字を削除
    name = re.sub(r'^\d+\s*', '', name)
    
    # キーワードを削除（前後にスペースまたは文字列の開始/終了がある場合のみ）
    # 長い順に削除（ロングドレスがドレスより先、キャバドレスがドレスより先）
    replacements = [
        (r'(?:^|\s)ロングドレス(?:\s|$)', ' '),
        (r'(?:^|\s)キャバドレス(?:\s|$)', ' '),
        (r'(?:^|\s)キャバクラ(?:\s|$)', ' '),
        (r'(?:^|\s)ドレス(?:\s|$)', ' '),  # キャミドレス、チャイナドレスは残る
    ]
    
    for pattern, replacement in replacements:
        name = re.sub(pattern, replacement, name)
    
    # 複数の空白を1つに
    name = re.sub(r'\s+', ' ', name)
    
    # 前後の空白を削除
    name = name.strip()
    
    return name

def get_latest_product_csv():
    """
    最新のproduct_data_*.csvファイルを取得
    """
    pattern = os.path.join(CSV_DIR, "product_data_*.csv")
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def update_excel_from_csv(completed_product_numbers, dry_run=False):
    """
    完了した品番に対して、CSVの情報でExcelを更新し、済列に日付も記入（1回の保存で完了）
    
    Args:
        completed_product_numbers: 完了した品番のリスト
        dry_run: ドライランモード
    
    Returns:
        更新成功した品番のリスト
    """
    if not completed_product_numbers:
        return []
    
    print(f"\n{'='*60}")
    print("📝 Excelファイルの更新処理を開始...")
    
    # CSVファイルを取得
    csv_path = get_latest_product_csv()
    if not csv_path:
        print("❌ product_data_*.csvが見つかりません")
        return []
    
    print(f"✅ CSVファイル: {os.path.basename(csv_path)}")
    
    # CSVを読み込む
    try:
        df_csv = pd.read_csv(csv_path, encoding='cp932')
        print(f"✅ CSV読み込み完了: {len(df_csv)}件")
    except Exception as e:
        print(f"❌ CSV読み込みエラー: {e}")
        return []
    
    # サイズと商品名1を生成
    print(f"⏳ サイズと商品名1を生成中...")
    df_csv['サイズ'] = df_csv['商品説明'].apply(extract_size_from_description)
    df_csv['商品名1'] = df_csv['商品名'].apply(clean_product_name)
    print(f"✅ サイズと商品名1を生成完了")
    
    # CSVから品番ごとのデータを抽出
    csv_data = {}
    for idx, row in df_csv.iterrows():
        # 商品名から品番を抽出
        product_from_name = extract_product_number_from_text(row.get('商品名', ''))
        product_from_desc = extract_product_number_from_text(row.get('商品説明', ''))
        
        # 完了品番リストにある場合のみ取得
        for product_number in [product_from_name, product_from_desc]:
            if product_number and product_number in completed_product_numbers:
                csv_data[product_number] = {
                    'brand_id': row.get('ブランドID', ''),
                    'name': row.get('商品名', ''),
                    'description': row.get('商品説明', ''),
                    'size': row.get('サイズ', 'F'),
                    'name1': row.get('商品名1', '')
                }
                break
    
    print(f"✅ CSVから {len(csv_data)} 件のデータを取得")
    
    if not csv_data:
        print("⚠️ CSVに該当する品番が見つかりませんでした")
        return []
    
    # Excelを読み込む
    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook[SHEET_NAME]
    except Exception as e:
        print(f"❌ Excel読み込みエラー: {e}")
        return []
    
    # Excelのヘッダーを探す
    header_row = None
    product_col = brand_id_col = name_col = desc_col = done_col = size_col = name1_col = None
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        for col_idx, cell in enumerate(row, start=1):
            value = str(cell.value or "").strip()
            if not product_col and value in ["品番", "商品番号", "ヒンバン", "hinban"]:
                product_col = col_idx
                header_row = row_idx
            if not brand_id_col and value in ["ブランドID", "brand_id", "brandid"]:
                brand_id_col = col_idx
            if not name_col and value in ["商品名", "product_name", "name"]:
                name_col = col_idx
            if not desc_col and value in ["商品説明", "description", "desc"]:
                desc_col = col_idx
            if not done_col and value in ["済", "済み", "完了", "Done"]:
                done_col = col_idx
            if not size_col and value in ["サイズ", "size", "SIZE"]:
                size_col = col_idx
            if not name1_col and value in ["商品名1", "商品名１", "name1"]:
                name1_col = col_idx
        
        if product_col:
            break
    
    if not product_col:
        print("❌ Excelの品番列が見つかりません")
        return []
    
    print(f"✅ Excel列: 品番={chr(64+product_col)}, ブランドID={chr(64+brand_id_col) if brand_id_col else 'なし'}, 商品名={chr(64+name_col) if name_col else 'なし'}, 商品説明={chr(64+desc_col) if desc_col else 'なし'}, サイズ={chr(64+size_col) if size_col else 'なし'}, 商品名1={chr(64+name1_col) if name1_col else 'なし'}, 済={chr(64+done_col) if done_col else 'なし'}")
    
    # 今日の日付（YYMMDD形式）
    today = datetime.now().strftime("%y%m%d")
    
    # 完了品番のセット（高速検索用）
    completed_set = set(completed_product_numbers)
    
    # Excelを更新（CSV更新 + 済列更新を同時に実行）
    updated_count = 0
    completed_count = 0
    updated_products = []
    data_start_row = (header_row or 1) + 1
    
    for row in sheet.iter_rows(min_row=data_start_row, values_only=False):
        product_value = row[product_col - 1].value if len(row) >= product_col else None
        if not product_value:
            continue
        
        product_number = str(product_value).strip().lstrip('0')
        
        # CSVデータで更新
        if product_number in csv_data:
            csv_info = csv_data[product_number]
            has_update = False
            
            # ブランドID更新
            if brand_id_col and len(row) >= brand_id_col and csv_info['brand_id']:
                row[brand_id_col - 1].value = csv_info['brand_id']
                has_update = True
            
            # 商品名更新
            if name_col and len(row) >= name_col and csv_info['name']:
                row[name_col - 1].value = csv_info['name']
                has_update = True
            
            # 商品説明更新
            if desc_col and len(row) >= desc_col and csv_info['description']:
                row[desc_col - 1].value = csv_info['description']
                has_update = True
            
            # サイズ更新
            if size_col and len(row) >= size_col and csv_info['size']:
                row[size_col - 1].value = csv_info['size']
                has_update = True
            
            # 商品名1更新
            if name1_col and len(row) >= name1_col and csv_info['name1']:
                row[name1_col - 1].value = csv_info['name1']
                has_update = True
            
            if has_update:
                print(f"  ✅ 品番 {product_number}: CSV更新成功")
                updated_count += 1
                updated_products.append(product_number)
        
        # 済列に日付を記入
        if done_col and product_number in completed_set:
            if len(row) >= done_col:
                row[done_col - 1].value = today
                print(f"  ✅ 品番 {product_number}: 済列に日付記入 ({today})")
                completed_count += 1
    
    print(f"\nCSV更新件数: {updated_count}件")
    if completed_count > 0:
        print(f"済列更新件数: {completed_count}件")
    
    # Excelを保存（1回だけ）
    total_changes = updated_count + completed_count
    if not dry_run and total_changes > 0:
        saved = False
        try:
            workbook.save(EXCEL_FILE)
            print(f"\n✅ Excelファイルを保存しました: 合計{total_changes}件更新")
            saved = True
        except PermissionError as e:
            # ファイルが開いている場合は別名で保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(EXCEL_FILE)[0]
            backup_file = f"{base_name}_更新_{timestamp}.xlsx"
            try:
                workbook.save(backup_file)
                print(f"\n⚠️ 元のファイルが開いているため、別名で保存しました:")
                print(f"   {os.path.basename(backup_file)}")
                print(f"   元のファイルを閉じた後、このファイルをリネームしてください")
                saved = True
            except Exception as e2:
                print(f"\n❌ 別名保存もエラー: {e2}")
                return []
        except Exception as e:
            print(f"\n❌ Excel保存エラー: {e}")
            return []
        
        if not saved:
            return []
    elif dry_run:
        print(f"\n🔍 [ドライラン] Excel更新予定: 合計{total_changes}件")
    
    print(f"{'='*60}")
    return updated_products

def read_product_numbers_from_excel():
    """
    Excelファイルから品番と明るさを読み取る
    条件：品番列に値があり、済列が空白
    
    Returns:
        (品番, 明るさ係数)のタプルのリスト
    """
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excelファイルが見つかりません: {EXCEL_FILE}")
        return []
    
    try:
        print(f"📄 Excelファイルを読み込み中: {os.path.basename(EXCEL_FILE)}")
        workbook = load_workbook(EXCEL_FILE, data_only=True)
        
        if SHEET_NAME not in workbook.sheetnames:
            print(f"❌ シート「{SHEET_NAME}」が見つかりません")
            print(f"   利用可能なシート: {', '.join(workbook.sheetnames)}")
            return []
        
        sheet = workbook[SHEET_NAME]
        print(f"✓ シート「{SHEET_NAME}」を読み込みました")
        
        # ヘッダー行を探す
        header_row = None
        product_col = None
        done_col = None
        brightness_col = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
            for col_idx, cell in enumerate(row, start=1):
                value = str(cell.value or "").strip()
                # 品番列を探す
                if not product_col and value in ["品番", "商品番号", "ヒンバン", "hinban"]:
                    product_col = col_idx
                    header_row = row_idx
                # 済列を探す
                if not done_col and value in ["済", "済み", "完了", "Done"]:
                    done_col = col_idx
                    header_row = row_idx
                # 明るさ列を探す
                if not brightness_col and value in ["明るさ", "明度", "brightness"]:
                    brightness_col = col_idx
                    header_row = row_idx
            
            if product_col and done_col:
                break
        
        if not product_col:
            print(f"❌ 品番列が見つかりません")
            return []
        
        if not done_col:
            print(f"⚠️ 済列が見つかりません（全ての品番を対象とします）")
        
        print(f"✓ 品番列: {chr(64 + product_col)}列")
        if done_col:
            print(f"✓ 済列: {chr(64 + done_col)}列")
        if brightness_col:
            print(f"✓ 明るさ列: {chr(64 + brightness_col)}列")
        
        # データ行を読み取る
        product_data = []
        data_start_row = (header_row or 1) + 1
        
        for row in sheet.iter_rows(min_row=data_start_row, values_only=False):
            # 品番列の値を取得
            product_value = row[product_col - 1].value if len(row) >= product_col else None
            
            if not product_value:
                continue
            
            product_number = str(product_value).strip()
            if not product_number:
                continue
            
            # 先頭の0を取り除く
            product_number = product_number.lstrip('0')
            if not product_number:  # 全て0だった場合はスキップ
                continue
            
            # 明るさ係数を取得（明るさ列に値がある場合のみ）
            brightness_factor = None
            if brightness_col:
                brightness_value = row[brightness_col - 1].value if len(row) >= brightness_col else None
                if brightness_value is not None:
                    try:
                        brightness_factor = float(brightness_value)
                    except (ValueError, TypeError):
                        pass
            
            # 済列のチェック
            if done_col:
                done_value = row[done_col - 1].value if len(row) >= done_col else None
                # 済列が空白またはNoneの場合のみ追加
                if done_value is None or str(done_value).strip() == "":
                    product_data.append((product_number, brightness_factor))
            else:
                # 済列がない場合は全て追加
                product_data.append((product_number, brightness_factor))
        
        print(f"✓ 抽出した品番: {len(product_data)}件")
        
        return product_data
        
    except Exception as e:
        print(f"❌ Excelファイル読み込みエラー: {e}")
        import traceback
        traceback.print_exc()
        return []

def process_products_with_custom_brightness(product_data, default_brightness, backup, copy_mode, dry_run):
    """
    品番ごとに異なる明るさ係数で処理
    
    Args:
        product_data: (品番, 明るさ係数)のタプルのリスト
        default_brightness: 明るさ列に値がない場合のデフォルト係数
        backup: バックアップを取るかどうか
        copy_mode: True=コピー（元を残す）, False=移動
        dry_run: ドライランモードかどうか
    """
    from brighten_images import move_images_to_re, brighten_image
    
    IMAGE_DIR = r"\\LS210DNBD82\share\平良\Python\mercari_dorekai\mercari_images"
    RE_DIR = r"\\LS210DNBD82\share\平良\Python\mercari_dorekai\mercari_images\re"
    
    if not os.path.exists(IMAGE_DIR):
        print(f"❌ 画像フォルダが見つかりません: {IMAGE_DIR}")
        return
    
    os.makedirs(RE_DIR, exist_ok=True)
    
    success_count = 0
    not_found_count = 0
    error_count = 0
    total_moved = 0
    skipped_count = 0
    
    # 明るさ係数別にグループ化
    brightness_groups = {}
    for product_number, brightness_factor in product_data:
        brightness = brightness_factor if brightness_factor is not None else default_brightness
        if brightness not in brightness_groups:
            brightness_groups[brightness] = []
        brightness_groups[brightness].append(product_number)
    
    print(f"\n{'='*60}")
    if dry_run:
        print("🔍 ドライランモード（プレビューのみ、実際の変更なし）")
    print(f"処理対象: {len(product_data)}件")
    print(f"デフォルト明るさ係数: {default_brightness}")
    print(f"移動モード: {'コピー（元を残す）' if copy_mode else '移動（元を削除）'}")
    print(f"バックアップ: {'あり' if backup else 'なし'}")
    print(f"移動先: {RE_DIR}")
    
    # 明るさ係数別の件数を表示
    print(f"\n明るさ係数別:")
    for brightness in sorted(brightness_groups.keys()):
        count = len(brightness_groups[brightness])
        print(f"  {brightness}: {count}件")
    
    print(f"{'='*60}\n")
    
    completed_products = []  # 成功した品番を記録
    
    for product_number, brightness_factor in product_data:
        brightness = brightness_factor if brightness_factor is not None else default_brightness
        
        print(f"▶ 品番: {product_number} (明るさ: {brightness})")
        
        # reフォルダに既に画像が存在するかチェック
        re_images = glob.glob(os.path.join(RE_DIR, f"{product_number}-*.jpg"))
        if re_images:
            print(f"    ℹ️ reフォルダに既に画像が存在（{len(re_images)}枚）- 画像処理スキップ")
            skipped_count += 1
            # Excel更新は実行するため、完了リストに追加
            completed_products.append(product_number)
            continue
        
        # 1. 全画像をreフォルダに移動/コピー
        moved_count, first_image_path = move_images_to_re(product_number, copy_mode, dry_run)
        
        if moved_count == 0:
            print(f"    ⚠️ 画像が見つかりません")
            not_found_count += 1
            continue
        
        total_moved += moved_count
        
        # 2. 1枚目のみ明るさ調整
        if first_image_path:
            if brighten_image(first_image_path, brightness, backup, dry_run):
                success_count += 1
                completed_products.append(product_number)  # 成功した品番を追加
            else:
                error_count += 1
        else:
            print(f"    ⚠️ 1枚目の画像が見つかりません")
            not_found_count += 1
    
    # 結果表示
    print(f"\n{'='*60}")
    if dry_run:
        print(f"🔍 ドライラン完了（実際の変更なし）")
        print(f"  移動/コピー予定: {total_moved}枚")
        print(f"  明るさ調整予定: {success_count}枚")
        print(f"  見つからない: {not_found_count}件")
        if skipped_count > 0:
            print(f"  スキップ: {skipped_count}件（reフォルダに既存）")
        print(f"\n💡 実際に実行する場合は --dry-run オプションなしで実行してください")
    else:
        print(f"✅ 画像処理完了")
        print(f"  移動/コピー成功: {total_moved}枚")
        print(f"  明るさ調整成功: {success_count}枚")
        print(f"  見つからない: {not_found_count}件")
        if skipped_count > 0:
            print(f"  スキップ: {skipped_count}件（reフォルダに既存）")
        print(f"  エラー: {error_count}枚")
        if backup and success_count > 0:
            print(f"\n💾 バックアップ保存先:")
            BACKUP_DIR = r"\\LS210DNBD82\share\平良\Python\mercari_dorekai\mercari_images\バックアップ_明るさ調整前"
            print(f"   {BACKUP_DIR}")
            print(f"   ※バックアップは1枚のみ保存されます（既存がある場合はスキップ）")
        print(f"\n📁 調整後の画像保存先:")
        print(f"   ・reフォルダ: {RE_DIR}")
        print(f"   ・元フォルダ: {IMAGE_DIR} (上書き)")
        if not copy_mode:
            print(f"\n⚠️ 元フォルダから画像を移動しました")
    print(f"{'='*60}")
    
    # 画像処理が成功した場合、ExcelをCSVの情報で更新（CSV更新 + 済列更新を1回で実行）
    if not dry_run and completed_products:
        # ExcelをCSVの情報で更新し、済列に日付も記入（1回の保存で完了）
        updated_products = update_excel_from_csv(completed_products, dry_run)
        
        if not updated_products:
            print("\n⚠️ Excelの更新に失敗しました")
    elif dry_run and completed_products:
        print(f"\n🔍 [ドライラン] Excel更新予定: {len(completed_products)}件")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Excelから品番を読み取り、画像をreフォルダに移動して1枚目を明るくする',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Excelファイル: {EXCEL_FILE}
対象シート: {SHEET_NAME}
条件: 品番列に値があり、済列が空白

使用例:
  # ドライラン（プレビュー）
  python brighten_from_excel.py --dry-run
  
  # 実行（コピーモード）
  python brighten_from_excel.py
  
  # 実行（移動モード）
  python brighten_from_excel.py --move
  
  # 明るさ調整（2倍明るく）
  python brighten_from_excel.py --brightness 2.0
        """
    )
    
    parser.add_argument('-b', '--brightness', type=float, default=1.2,
                        help='明るさの係数（デフォルト: 1.2、Excelの明るさ列の値がない場合に使用）')
    parser.add_argument('--move', action='store_true',
                        help='移動モード（元フォルダから削除、デフォルトはコピー）')
    parser.add_argument('--no-backup', action='store_true',
                        help='バックアップを作成しない')
    parser.add_argument('--dry-run', action='store_true',
                        help='ドライランモード（プレビューのみ、実際の変更なし）')
    
    args = parser.parse_args()
    
    # Excelから品番と明るさを読み取る
    product_data = read_product_numbers_from_excel()
    
    if not product_data:
        print("\n❌ 処理対象の品番がありません")
        return
    
    print(f"\n📋 処理対象の品番:")
    for i, (pn, brightness) in enumerate(product_data[:10], 1):
        brightness_str = f" (明るさ: {brightness})" if brightness is not None else ""
        print(f"  {i}. {pn}{brightness_str}")
    if len(product_data) > 10:
        print(f"  ... 他 {len(product_data) - 10}件")
    
    # 確認
    if not args.dry_run:
        print(f"\n⚠️ {len(product_data)}件の品番を処理します")
        response = input("続行しますか？ (Y/n): ").strip().lower()
        if response in ['n', 'no']:
            print("キャンセルしました")
            return
    
    # 処理実行
    process_products_with_custom_brightness(
        product_data=product_data,
        default_brightness=args.brightness,
        backup=not args.no_backup,
        copy_mode=not args.move,
        dry_run=args.dry_run
    )

if __name__ == "__main__":
    main()
