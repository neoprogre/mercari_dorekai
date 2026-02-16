from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\progr\Desktop\Python\mercari_dorekai\downloads\ドレ買い.xlsx', data_only=True)

print('シート一覧:')
print(wb.sheetnames)

if 'シート1' in wb.sheetnames:
    sheet1 = wb['シート1']
    
    # 3行目から確認（ヘッダーがある行）
    print('\n3行目（ヘッダー）の全列:')
    row3 = []
    for col_idx, cell in enumerate(sheet1[3], start=1):
        if cell.value:
            row3.append(f'{col_idx}:{cell.value}')
    print(row3)
    
    # 4行目のデータ（最初のデータ行）
    print('\n4行目のデータ:')
    row4_values = []
    for col_idx, cell in enumerate(sheet1[4], start=1):
        if cell.value:
            row4_values.append(f'{col_idx}:{cell.value}')
    print(row4_values[:20])  # 最初の20列
else:
    print('\n❌ シート1が見つかりません')

